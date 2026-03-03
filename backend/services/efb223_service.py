"""
EFB 223 Service — consolidated from efb223-generator app.
Parses PDFs via Claude API, generates Excel + PDF files.
"""
import anthropic
import base64
import io
import json
import math
import os
import re
import logging
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from fpdf import FPDF
import pikepdf

logger = logging.getLogger(__name__)

# ── Constants ────────────────────────────────────────────────────────
HOURLY_RATE = 71.10
LABOR_RATIO = 0.52
STOFFE_RATIO = 0.46

# ── Excel Styles ─────────────────────────────────────────────────────
thin = Side(style="thin")
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

tr10 = Font(name="Times New Roman", size=10)
ar8 = Font(name="Arial", size=8)
ar6 = Font(name="Arial", size=6)

left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
center_top = Alignment(horizontal="center", vertical="top", wrap_text=True)
right_top = Alignment(horizontal="right", vertical="top", wrap_text=True)
right_bottom = Alignment(horizontal="right", vertical="bottom", wrap_text=True)
left_center = Alignment(horizontal="left", vertical="center", wrap_text=True)

WIDTHS = {"A": 10, "B": 26, "C": 10, "D": 9, "E": 13, "F": 8, "G": 6, "H": 12, "I": 12, "J": 4, "K": 8, "L": 15}

# ── PDF Constants ────────────────────────────────────────────────────
PDF_TOTAL_W = 178
PDF_MARGIN = 10
PDF_ACTIVE_COLS = [1, 2, 3, 4, 5, 6, 8, 9, 10, 12]
PDF_COL_W = {1: 14, 2: 38, 3: 14, 4: 12, 5: 16, 6: 18, 8: 16, 9: 16, 10: 14, 12: 20}
PDF_HEADERS = {
    1: "OZ\ndes LV 1", 2: "Kurzbezeichnung d.\nTeilleistung 1", 3: "Menge 1",
    4: "Mengen-\neinheit 1", 5: "Zeitan-\nsatz 2", 6: "L\u00f6hne 2, 3",
    8: "Stoffe 2", 9: "Ger\u00e4te 2, 4", 10: "Sonstig. 2",
    12: "Angeboten.\nEinheitspreis\n(Sp. 6+7+8+9)",
}
PDF_NUMS = {1: "1", 2: "2", 3: "3", 4: "4", 5: "5", 6: "6", 8: "7", 9: "8", 10: "9", 12: "10"}

# ── Claude API System Prompt ─────────────────────────────────────────
SYSTEM_PROMPT = """You are an expert construction document parser for German EFB 223 (Einheitliche Formbl\u00e4tter) documents.

You will receive two PDF documents:
1. An EFB 223 template PDF - contains OZ numbers, Menge (quantity), Einheit (unit)
2. A Leistungsverzeichnis (LV) PDF - contains full position descriptions with prices

Your task: Extract ALL positions and return structured JSON.

RULES:
- Extract EVERY position from the LV, cross-referencing with the EFB 223 template
- OZ numbers must match exactly as shown in the documents (e.g. "1.1.10.", "2.1.10.")
- Menge and Einheit come from the EFB 223 template
- Einheitspreis (EP) and Gesamtpreis (GP) come from the LV
- GP = Menge x EP (verify this)
- Section headers (like "1", "1.1.", "2", "2.1.") are non-data rows with no prices
- Kurzbezeichnung should be a concise description (not the full LV text)

For each data position, classify the Loehne/Stoffe/Geraete percentage split based on the work type:
- Concrete precast elements (Betonfertigteile, Wandplatte, Balkonbruestung): L=52, S=45, G=3
- Vertical elements with anchors (Haengezuganker): L=50, S=47, G=3
- Reinforcement steel (Betonstabstahl, Betonstahlmatte): L=35, S=65, G=0
- Joint sealing (Bauteilfuge abdichten): L=50, S=50, G=0
- Sand coating supplement (Zulage Besandung): L=60, S=40, G=0
- Steel pipes/fittings (Rohr, Bogen, Sprungrohr, Muffe, Abzweig): L=40, S=60, G=0
- Steel construction (Einbauteile BFT): L=50, S=48, G=2
- Steel construction support (Auflager, Konsolen, HEB): L=45, S=53, G=2 (except large konsolen: L=43, S=55, G=2)
- Anchor rails (Ankerschiene): L=45, S=52, G=3
- Threaded rod (Gewindestab): L=40, S=60, G=0
- Tension/compression system (Zug-Druckstabsystem): L=45, S=53, G=2
- Insulation (Perimeterdaemmung XPS): L=45, S=55, G=0
- Insulation (Waermedaemmung Mineralwolle): L=48, S=52, G=0
- Additional insulation work (Mehraufwand Daemmplatten): L=85, S=15, G=0
- Waterproofing (Abdichtung EPDM/FLK): L=45, S=55, G=0
- Waterproofing balcony (Abdichtung 2K-PUR): L=48, S=52, G=0
- Scaffolding (Arbeitsgeruest): L=55, S=40, G=5
- Electrical conduit (Elektroinstallationsrohr): L=45, S=55, G=0
- Electrical box (Grossrohrdose): L=40, S=60, G=0
- Drain (Einzelablauf): L=40, S=60, G=0
- Water spout (Wasserspeier): L=40, S=60, G=0
- Pattern/sample (Muster, Bemusterung): L=50, S=50, G=0
- Graffiti protection (Graffitischutz): L=30, S=70, G=0
- Site setup (Baustelleneinrichtung, Einrichten): L=70, S=25, G=5
- Site maintenance (vorhalten): L=60, S=35, G=5
- Corner protection (Eckschutz): L=50, S=50, G=0
- Technical work/planning (Technische Bearbeitung, Planung, Dokumentation, Koordination, BNB, Ingenieur, Polier, Facharbeiter, Helfer, Stundenlohn): L=100, S=0, G=0

You MUST return ONLY valid JSON. No markdown, no code fences, no comments, no trailing commas.

The JSON structure:
{"project":{"vergabenummer":"string","baumasnahme":"string","leistung":"string"},"positions":[{"oz":"1","beschreibung":"Section Name","menge":null,"einheit":null,"ep":null,"gp":null,"loehne_pct":null,"stoffe_pct":null,"geraete_pct":null,"is_section":true},{"oz":"1.1.10.","beschreibung":"Short description","menge":1.0,"einheit":"St","ep":500.00,"gp":500.00,"loehne_pct":70,"stoffe_pct":25,"geraete_pct":5,"is_section":false}]}

CRITICAL RULES:
- All prices in EUR with decimal points (not commas)
- Menge can be decimal (e.g. 100.600, 3.800)
- Section headers: is_section=true, all numeric fields null
- Percentages must sum to 100 for each data position
- Output ONLY the raw JSON object, nothing else"""


# ══════════════════════════════════════════════════════════════
#  PDF PARSING (Claude API)
# ══════════════════════════════════════════════════════════════

def _repair_json(raw: str) -> str:
    """Attempt to repair common JSON issues from LLM output."""
    text = raw.strip()
    if text.startswith("```"):
        text = re.sub(r'^```[a-zA-Z]*\n?', '', text)
    if text.endswith("```"):
        text = text[:text.rfind("```")]
    text = text.strip()

    start = text.find("{")
    if start == -1:
        raise ValueError("No JSON object found in response")

    depth = 0
    end = -1
    in_string = False
    escape = False
    for i in range(start, len(text)):
        ch = text[i]
        if escape:
            escape = False
            continue
        if ch == '\\' and in_string:
            escape = True
            continue
        if ch == '"' and not escape:
            in_string = not in_string
            continue
        if in_string:
            continue
        if ch == '{':
            depth += 1
        elif ch == '}':
            depth -= 1
            if depth == 0:
                end = i
                break

    if end != -1:
        text = text[start:end + 1]
    else:
        text = text[start:]
        open_brackets = 0
        open_braces = 0
        in_str = False
        esc = False
        for ch in text:
            if esc:
                esc = False
                continue
            if ch == '\\' and in_str:
                esc = True
                continue
            if ch == '"':
                in_str = not in_str
                continue
            if in_str:
                continue
            if ch == '[':
                open_brackets += 1
            elif ch == ']':
                open_brackets -= 1
            elif ch == '{':
                open_braces += 1
            elif ch == '}':
                open_braces -= 1
        text = re.sub(r',\s*$', '', text)
        text += ']' * open_brackets
        text += '}' * open_braces

    text = re.sub(r',\s*([}\]])', r'\1', text)
    text = re.sub(r'//[^\n]*', '', text)
    return text


def parse_pdfs(efb_pdf_bytes: bytes, lv_pdf_bytes: bytes, api_key: str) -> dict:
    """Send both PDFs to Claude API and extract structured position data."""
    client = anthropic.Anthropic(api_key=api_key)
    efb_b64 = base64.standard_b64encode(efb_pdf_bytes).decode("utf-8")
    lv_b64 = base64.standard_b64encode(lv_pdf_bytes).decode("utf-8")

    logger.info("Sending PDFs to Claude API (streaming)...")
    collected_text = []
    stop_reason = None

    with client.messages.stream(
        model="claude-sonnet-4-20250514",
        max_tokens=64000,
        system=SYSTEM_PROMPT,
        messages=[{
            "role": "user",
            "content": [
                {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": efb_b64}, "cache_control": {"type": "ephemeral"}},
                {"type": "text", "text": "Above is the EFB 223 template PDF."},
                {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": lv_b64}, "cache_control": {"type": "ephemeral"}},
                {"type": "text", "text": "Above is the Leistungsverzeichnis (LV) PDF. Extract ALL positions from both documents and return the structured JSON as specified. Output ONLY valid JSON, no other text."},
            ],
        }],
    ) as stream:
        for text in stream.text_stream:
            collected_text.append(text)
        response = stream.get_final_message()
        stop_reason = response.stop_reason

    raw = "".join(collected_text)
    logger.info(f"Claude response: {len(raw)} chars, stop_reason={stop_reason}")
    if stop_reason == "max_tokens":
        logger.warning("Response was truncated (hit max_tokens). Will attempt JSON repair.")

    repaired = _repair_json(raw)
    try:
        data = json.loads(repaired)
    except json.JSONDecodeError as e:
        logger.error(f"JSON parse failed after repair: {e}")
        raise ValueError(f"Failed to parse Claude response as JSON: {e}")

    if "positions" not in data:
        raise ValueError("Claude response missing 'positions' key")

    valid_positions = []
    for pos in data["positions"]:
        if pos.get("is_section"):
            valid_positions.append(pos)
        else:
            try:
                for field in ("menge", "ep", "gp"):
                    if pos.get(field) is not None:
                        pos[field] = float(pos[field])
                for field in ("loehne_pct", "stoffe_pct", "geraete_pct"):
                    if pos.get(field) is not None:
                        pos[field] = int(pos[field])
                pct_sum = (pos.get("loehne_pct") or 0) + (pos.get("stoffe_pct") or 0) + (pos.get("geraete_pct") or 0)
                if pct_sum != 100 and pct_sum > 0:
                    pos["loehne_pct"] = 100 - (pos.get("stoffe_pct") or 0) - (pos.get("geraete_pct") or 0)
                valid_positions.append(pos)
            except (ValueError, TypeError) as e:
                logger.warning(f"Skipping invalid position {pos.get('oz')}: {e}")

    data["positions"] = valid_positions
    logger.info(f"Parsed {len(valid_positions)} positions ({sum(1 for p in valid_positions if not p.get('is_section'))} data rows)")
    return data


# ══════════════════════════════════════════════════════════════
#  EXCEL GENERATION
# ══════════════════════════════════════════════════════════════

def _apply_border(ws, row, c1, c2):
    for c in range(c1, c2 + 1):
        ws.cell(row=row, column=c).border = thin_border

def _set(ws, row, col, value, font=tr10, align=left_top):
    cell = ws.cell(row=row, column=col)
    cell.value = value
    cell.font = font
    cell.alignment = align
    cell.border = thin_border
    return cell

def _set_widths(ws):
    for col, w in WIDTHS.items():
        ws.column_dimensions[col].width = w

def _write_efb_header(ws, start_row, project, bieter, datum):
    r = start_row
    ws.merge_cells(f"A{r}:L{r}")
    _set(ws, r, 1, "223\n(Aufgliederung der Einheitspreise)", tr10, right_bottom)
    _apply_border(ws, r, 1, 12)
    ws.row_dimensions[r].height = 128.25
    r += 1

    ws.merge_cells(f"A{r}:F{r}")
    _set(ws, r, 1, f"Bieter\n{bieter}", Font(name="Arial", size=10), left_top)
    ws.merge_cells(f"G{r}:J{r}")
    _set(ws, r, 7, f"Vergabenummer\n{project.get('vergabenummer', '')}", tr10, left_top)
    ws.merge_cells(f"K{r}:L{r}")
    _set(ws, r, 11, f"Datum\n{datum}", Font(name="Arial", size=10), left_top)
    _apply_border(ws, r, 1, 12)
    ws.row_dimensions[r].height = 30
    r += 1

    ws.merge_cells(f"A{r}:L{r}")
    _set(ws, r, 1, f"Bauma\u00dfnahme {project.get('baumasnahme', '')}", tr10, left_top)
    _apply_border(ws, r, 1, 12)
    ws.row_dimensions[r].height = 39
    r += 1

    ws.merge_cells(f"A{r}:L{r}")
    _set(ws, r, 1, f"Angebot f\u00fcr\n{project.get('leistung', '')}", tr10, left_top)
    _apply_border(ws, r, 1, 12)
    ws.row_dimensions[r].height = 30
    r += 1

    ws.merge_cells(f"A{r}:L{r}")
    _set(ws, r, 1, "Aufgliederung der Einheitspreise", tr10, left_top)
    _apply_border(ws, r, 1, 12)
    ws.row_dimensions[r].height = 32
    r += 1

    ws.merge_cells(f"F{r}:L{r}")
    _set(ws, r, 6, "Teilkosten einschl. Zuschl\u00e4ge in EUR (ohne Umsatzsteuer) je Mengeneinheit 2", tr10, left_top)
    _apply_border(ws, r, 1, 12)
    ws.row_dimensions[r].height = 21
    r += 1

    _set(ws, r, 1, "OZ\ndes LV 1", tr10, center_top)
    _set(ws, r, 2, "Kurzbezeichnung d.\nTeilleistung 1", tr10, center_top)
    _set(ws, r, 3, "Menge 1", tr10, left_top)
    _set(ws, r, 4, "Mengen\n-einheit\n1", tr10, center_top)
    _set(ws, r, 5, "Zeitan- satz 2", tr10, left_top)
    ws.merge_cells(f"F{r}:G{r}")
    _set(ws, r, 6, "L\u00f6hne 2, 3", tr10, left_top)
    _set(ws, r, 8, "Stoffe  2", tr10, center_top)
    _set(ws, r, 9, "Ger\u00e4te 2, 4", tr10, center_top)
    _set(ws, r, 10, "Sonstig. 2", tr10, left_top)
    ws.merge_cells(f"J{r}:K{r}")
    _set(ws, r, 12, "Angeboten.\nEinheitspreis\n(Sp. 6+7+8+9)", tr10, left_top)
    _apply_border(ws, r, 1, 12)
    ws.row_dimensions[r].height = 52
    r += 1

    col_num_map = {1: 1, 2: 2, 3: 3, 4: 4, 5: 5, 6: 6, 8: 7, 9: 8, 10: 9, 12: 10}
    for col, num in col_num_map.items():
        cell = ws.cell(row=r, column=col)
        cell.value = num
        cell.font = ar6
        cell.alignment = Alignment(horizontal="center", vertical="top")
        cell.border = thin_border
    ws.merge_cells(f"F{r}:G{r}")
    ws.merge_cells(f"J{r}:K{r}")
    _apply_border(ws, r, 1, 12)
    ws.row_dimensions[r].height = 12.75
    r += 1
    return r

def _write_urkal_header(ws):
    ws.merge_cells("A1:L1")
    cell = ws.cell(row=1, column=1)
    cell.value = "Urkalkulation"
    cell.font = Font(name="Times New Roman", size=14, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
    for c in range(2, 13):
        ws.cell(row=1, column=c).border = thin_border
    ws.row_dimensions[1].height = 36

    r = 2
    _set(ws, r, 1, "OZ\ndes LV 1", tr10, center_top)
    _set(ws, r, 2, "Kurzbezeichnung d.\nTeilleistung 1", tr10, center_top)
    _set(ws, r, 3, "Menge 1", tr10, left_top)
    _set(ws, r, 4, "Mengen\n-einheit\n1", tr10, center_top)
    _set(ws, r, 5, "Zeitan- satz 2", tr10, left_top)
    ws.merge_cells(f"F{r}:G{r}")
    _set(ws, r, 6, "L\u00f6hne 2, 3", tr10, left_top)
    _set(ws, r, 8, "Stoffe  2", tr10, center_top)
    _set(ws, r, 9, "Ger\u00e4te 2, 4", tr10, center_top)
    _set(ws, r, 10, "Sonstig. 2", tr10, left_top)
    ws.merge_cells(f"J{r}:K{r}")
    _set(ws, r, 12, "Angeboten.\nEinheitspreis\n(Sp. 6+7+8+9)", tr10, left_top)
    _apply_border(ws, r, 1, 12)
    ws.row_dimensions[r].height = 52

    r = 3
    col_num_map = {1: 1, 2: 2, 3: 3, 4: 4, 5: 5, 6: 6, 8: 7, 9: 8, 10: 9, 12: 10}
    for col, num in col_num_map.items():
        cell = ws.cell(row=r, column=col)
        cell.value = num
        cell.font = ar6
        cell.alignment = Alignment(horizontal="center", vertical="top")
        cell.border = thin_border
    ws.merge_cells(f"F{r}:G{r}")
    ws.merge_cells(f"J{r}:K{r}")
    _apply_border(ws, r, 1, 12)
    ws.row_dimensions[r].height = 12.75
    return 4

def _write_data_row(ws, r, oz, beschr, menge, einheit, loehne, stoffe, geraete, ep, zeitansatz=None):
    _set(ws, r, 1, oz, ar8, left_top)
    _set(ws, r, 2, beschr, tr10, left_top)
    if menge is not None:
        _set(ws, r, 3, menge, ar8, right_top)
    _set(ws, r, 4, einheit, ar8, right_top)
    if zeitansatz is not None:
        _set(ws, r, 5, round(zeitansatz, 2), tr10, right_top)
    else:
        ws.cell(row=r, column=5).border = thin_border
    ws.merge_cells(f"F{r}:G{r}")
    _set(ws, r, 6, loehne, tr10, left_center if menge and menge <= 5 else left_top)
    _set(ws, r, 8, stoffe, tr10, left_center if menge and menge <= 5 else left_top)
    _set(ws, r, 9, geraete, tr10, left_center if menge and menge <= 5 else left_top)
    ws.merge_cells(f"J{r}:K{r}")
    ws.cell(row=r, column=10).border = thin_border
    _set(ws, r, 12, ep, tr10, left_center if menge and menge <= 5 else left_top)
    _apply_border(ws, r, 1, 12)
    if beschr:
        lines = max(1, len(beschr) // 32 + 1)
        ws.row_dimensions[r].height = max(16.5, lines * 14)
    else:
        ws.row_dimensions[r].height = 16.5

def _write_section_row(ws, r, oz, beschr):
    _set(ws, r, 1, oz, ar8, left_top)
    _set(ws, r, 2, beschr, ar8 if not oz.replace(".", "").isdigit() or "." in oz else tr10, left_top)
    ws.merge_cells(f"C{r}:L{r}")
    _apply_border(ws, r, 1, 12)
    ws.row_dimensions[r].height = 16.5

def _apply_discount(positions, nachlass_pct):
    factor = 1.0 + (nachlass_pct / 100.0)
    result = []
    for pos in positions:
        if pos.get("is_section"):
            result.append(dict(pos))
        else:
            new = dict(pos)
            new["ep"] = round(pos["ep"] * factor, 2)
            new["gp"] = round(new["ep"] * pos["menge"], 2)
            result.append(new)
    return result

def _write_positions(ws, start_row, positions, total_gp, total_hours):
    r = start_row
    for pos in positions:
        oz = pos["oz"]
        beschr = pos["beschreibung"]
        if pos.get("is_section"):
            _write_section_row(ws, r, oz, beschr)
            r += 1
        else:
            menge = pos["menge"]
            einheit = pos.get("einheit", "")
            ep = pos["ep"]
            gp = pos["gp"]
            l_pct = pos["loehne_pct"]
            s_pct = pos["stoffe_pct"]
            g_pct = pos["geraete_pct"]
            loehne_pu = round(ep * l_pct / 100, 2)
            stoffe_pu = round(ep * s_pct / 100, 2)
            geraete_pu = round(ep * g_pct / 100, 2)
            diff = round(ep - (loehne_pu + stoffe_pu + geraete_pu), 2)
            loehne_pu = round(loehne_pu + diff, 2)
            pos_share = gp / total_gp if total_gp else 0
            zeitansatz = pos_share * total_hours
            _write_data_row(ws, r, oz, beschr, menge, einheit, loehne_pu, stoffe_pu, geraete_pu, ep, zeitansatz)
            r += 1
    return r

def _generate_efb223(positions, project, bieter, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EFB 223"
    _set_widths(ws)
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    datum = datetime.now().strftime("%d.%m.%Y")
    current_row = _write_efb_header(ws, 1, project, bieter, datum)
    total_gp = sum(p["gp"] for p in positions if not p.get("is_section"))
    total_hours = total_gp * LABOR_RATIO / HOURLY_RATE
    _write_positions(ws, current_row, positions, total_gp, total_hours)
    wb.save(output_path)
    return total_gp

def _generate_urkalkulation(positions, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Urkalkulation"
    _set_widths(ws)
    current_row = _write_urkal_header(ws)
    total_gp = sum(p["gp"] for p in positions if not p.get("is_section"))
    total_hours = total_gp * LABOR_RATIO / HOURLY_RATE
    _write_positions(ws, current_row, positions, total_gp, total_hours)
    wb.save(output_path)
    return total_gp

def generate_all_excel(positions, project, bieter, nachlass_pct, output_dir):
    os.makedirs(output_dir, exist_ok=True)
    files = {}
    efb_path = os.path.join(output_dir, "EFB_223_Original.xlsx")
    urkal_path = os.path.join(output_dir, "Urkalkulation_Original.xlsx")
    total_orig = _generate_efb223(positions, project, bieter, efb_path)
    _generate_urkalkulation(positions, urkal_path)
    files["efb_xlsx"] = efb_path
    files["urkal_xlsx"] = urkal_path
    files["total_original"] = total_orig

    if nachlass_pct != 0:
        disc_positions = _apply_discount(positions, nachlass_pct)
        efb_n_path = os.path.join(output_dir, f"EFB_223_Nachlass_{abs(int(nachlass_pct))}.xlsx")
        urkal_n_path = os.path.join(output_dir, f"Urkalkulation_Nachlass_{abs(int(nachlass_pct))}.xlsx")
        total_nachlass = _generate_efb223(disc_positions, project, bieter, efb_n_path)
        _generate_urkalkulation(disc_positions, urkal_n_path)
        files["efb_nachlass_xlsx"] = efb_n_path
        files["urkal_nachlass_xlsx"] = urkal_n_path
        files["total_nachlass"] = total_nachlass
    return files


# ══════════════════════════════════════════════════════════════
#  PDF GENERATION (from Excel)
# ══════════════════════════════════════════════════════════════

def _cv(ws, row, col):
    val = ws.cell(row=row, column=col).value
    if val is None:
        return ""
    if isinstance(val, float):
        if val == int(val):
            return str(int(val))
        return f"{val:.2f}"
    return str(val)

def _col_x(c):
    return PDF_MARGIN + sum(PDF_COL_W[cc] for cc in PDF_ACTIVE_COLS[:PDF_ACTIVE_COLS.index(c)])

def _write_pdf_column_headers(pdf):
    pdf.set_font("Helvetica", "", 7)
    header_h = 14
    y = pdf.get_y()
    x = PDF_MARGIN
    for c in PDF_ACTIVE_COLS:
        pdf.rect(x, y, PDF_COL_W[c], header_h)
        x += PDF_COL_W[c]
    for c in PDF_ACTIVE_COLS:
        x = _col_x(c)
        pdf.set_xy(x + 0.5, y + 0.5)
        pdf.multi_cell(PDF_COL_W[c] - 1, 3.2, PDF_HEADERS[c], border=0, align="C")
    pdf.set_y(y + header_h)

def _write_pdf_column_numbers(pdf):
    pdf.set_font("Helvetica", "", 6)
    for c in PDF_ACTIVE_COLS:
        pdf.cell(PDF_COL_W[c], 4, PDF_NUMS[c], border=1, align="C")
    pdf.ln()

def _write_pdf_data_rows(pdf, ws, start_row):
    pdf.set_font("Helvetica", "", 8)
    row_num = start_row
    max_row = ws.max_row
    while row_num <= max_row:
        oz = _cv(ws, row_num, 1)
        beschr = _cv(ws, row_num, 2)
        menge = _cv(ws, row_num, 3)
        einheit = _cv(ws, row_num, 4)
        zeitansatz = _cv(ws, row_num, 5)
        loehne = _cv(ws, row_num, 6)
        stoffe = _cv(ws, row_num, 8)
        geraete = _cv(ws, row_num, 9)
        ep = _cv(ws, row_num, 12)
        is_section = menge == "" and loehne == "" and ep == ""

        if is_section:
            row_h = 5.5
            if pdf.get_y() + row_h > 285:
                pdf.add_page()
            pdf.set_font("Helvetica", "", 8)
            pdf.cell(PDF_COL_W[1], row_h, oz, border=1, align="L")
            pdf.cell(PDF_TOTAL_W - PDF_COL_W[1], row_h, beschr, border=1, align="L")
            pdf.ln()
        else:
            pdf.set_font("Helvetica", "", 8)
            chars_per_line = max(1, int(PDF_COL_W[2] / 1.8))
            lines = max(1, (len(beschr) + chars_per_line - 1) // chars_per_line)
            row_h = max(5.5, lines * 4)
            if pdf.get_y() + row_h > 285:
                pdf.add_page()
            y_start = pdf.get_y()
            x = PDF_MARGIN
            for c in PDF_ACTIVE_COLS:
                pdf.rect(x, y_start, PDF_COL_W[c], row_h)
                x += PDF_COL_W[c]
            vals = {1: oz, 2: beschr, 3: menge, 4: einheit, 5: zeitansatz, 6: loehne, 8: stoffe, 9: geraete, 10: "", 12: ep}
            for c in PDF_ACTIVE_COLS:
                x = _col_x(c)
                pdf.set_xy(x + 0.5, y_start + 0.5)
                if c == 2:
                    pdf.multi_cell(PDF_COL_W[c] - 1, 3.5, vals[c], border=0, align="L")
                else:
                    align = "R" if c in (3, 5, 6, 8, 9, 12) else "L"
                    pdf.cell(PDF_COL_W[c] - 1, 4, vals[c], border=0, align=align)
            pdf.set_y(y_start + row_h)
        row_num += 1

def _efb_to_pdf(xlsx_path, pdf_path, project):
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()
    pdf.set_left_margin(PDF_MARGIN)
    pdf.set_right_margin(PDF_MARGIN)

    pdf.set_font("Helvetica", "", 11)
    y = pdf.get_y()
    pdf.rect(PDF_MARGIN, y, PDF_TOTAL_W, 28)
    pdf.set_xy(PDF_MARGIN + 2, y + 14)
    pdf.cell(PDF_TOTAL_W - 4, 6, "223", align="R")
    pdf.set_xy(PDF_MARGIN + 2, y + 20)
    pdf.cell(PDF_TOTAL_W - 4, 6, "(Aufgliederung der Einheitspreise)", align="R")
    pdf.set_y(y + 28)

    y = pdf.get_y()
    row_h = 10
    w_bieter = sum(PDF_COL_W[c] for c in [1, 2, 3, 4, 5, 6])
    pdf.rect(PDF_MARGIN, y, w_bieter, row_h)
    pdf.set_xy(PDF_MARGIN + 1, y + 0.5)
    pdf.set_font("Helvetica", "", 7)
    pdf.cell(30, 3.5, "Bieter")
    pdf.set_xy(PDF_MARGIN + 1, y + 4)
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(w_bieter - 2, 5, project.get("bieter", ""))

    w_verg = PDF_COL_W[8] + PDF_COL_W[9] + PDF_COL_W[10]
    x_verg = PDF_MARGIN + w_bieter
    pdf.rect(x_verg, y, w_verg, row_h)
    pdf.set_xy(x_verg + 1, y + 0.5)
    pdf.set_font("Helvetica", "", 7)
    pdf.cell(30, 3.5, "Vergabenummer")
    pdf.set_xy(x_verg + 1, y + 4)
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(w_verg - 2, 5, project.get("vergabenummer", ""))

    w_datum = PDF_TOTAL_W - w_bieter - w_verg
    x_datum = x_verg + w_verg
    pdf.rect(x_datum, y, w_datum, row_h)
    pdf.set_xy(x_datum + 1, y + 0.5)
    pdf.set_font("Helvetica", "", 7)
    pdf.cell(20, 3.5, "Datum")
    pdf.set_xy(x_datum + 1, y + 4)
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(w_datum - 2, 5, project.get("datum", ""))
    pdf.set_y(y + row_h)

    y = pdf.get_y()
    pdf.rect(PDF_MARGIN, y, PDF_TOTAL_W, row_h)
    pdf.set_xy(PDF_MARGIN + 1, y + 0.5)
    pdf.set_font("Helvetica", "", 7)
    pdf.cell(30, 3.5, "Bauma\u00dfnahme")
    pdf.set_xy(PDF_MARGIN + 1, y + 4)
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(PDF_TOTAL_W - 2, 5, project.get("baumasnahme", ""))
    pdf.set_y(y + row_h)

    y = pdf.get_y()
    pdf.rect(PDF_MARGIN, y, PDF_TOTAL_W, row_h)
    pdf.set_xy(PDF_MARGIN + 1, y + 0.5)
    pdf.set_font("Helvetica", "", 7)
    pdf.cell(30, 3.5, "Angebot f\u00fcr")
    pdf.set_xy(PDF_MARGIN + 1, y + 4)
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(PDF_TOTAL_W - 2, 5, project.get("leistung", ""))
    pdf.set_y(y + row_h)

    y = pdf.get_y()
    row_h2 = 8
    pdf.rect(PDF_MARGIN, y, PDF_TOTAL_W, row_h2)
    pdf.set_xy(PDF_MARGIN + 1, y + 1.5)
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(PDF_TOTAL_W - 2, 5, "Aufgliederung der Einheitspreise")
    pdf.set_y(y + row_h2)

    y = pdf.get_y()
    row_h3 = 6
    w_ae = sum(PDF_COL_W[c] for c in [1, 2, 3, 4, 5])
    pdf.rect(PDF_MARGIN, y, w_ae, row_h3)
    w_fl = PDF_TOTAL_W - w_ae
    pdf.rect(PDF_MARGIN + w_ae, y, w_fl, row_h3)
    pdf.set_xy(PDF_MARGIN + w_ae + 1, y + 0.5)
    pdf.set_font("Helvetica", "", 7)
    pdf.cell(w_fl - 2, 4, "Teilkosten einschl. Zuschl\u00e4ge in EUR (ohne Umsatzsteuer) je Mengeneinheit 2")
    pdf.set_y(y + row_h3)

    _write_pdf_column_headers(pdf)
    _write_pdf_column_numbers(pdf)
    _write_pdf_data_rows(pdf, ws, 9)
    pdf.output(pdf_path)

def _urkal_to_pdf(xlsx_path, pdf_path, password=None):
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()
    pdf.set_left_margin(PDF_MARGIN)
    pdf.set_right_margin(PDF_MARGIN)
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(PDF_TOTAL_W, 10, "Urkalkulation", border=1, ln=1, align="C")
    pdf.ln(1)
    _write_pdf_column_headers(pdf)
    _write_pdf_column_numbers(pdf)
    _write_pdf_data_rows(pdf, ws, 4)

    if password:
        temp_pdf = pdf_path + ".tmp"
        pdf.output(temp_pdf)
        with pikepdf.open(temp_pdf) as p:
            p.save(pdf_path, encryption=pikepdf.Encryption(owner=password, user=password, R=6))
        os.remove(temp_pdf)
    else:
        pdf.output(pdf_path)

def generate_all_pdfs(excel_files, project, bieter, password=None):
    pdf_project = dict(project)
    pdf_project["bieter"] = bieter
    pdf_project["datum"] = datetime.now().strftime("%d.%m.%Y")
    pdfs = {}

    efb_pdf = excel_files["efb_xlsx"].replace(".xlsx", ".pdf")
    _efb_to_pdf(excel_files["efb_xlsx"], efb_pdf, pdf_project)
    pdfs["efb_pdf"] = efb_pdf

    urkal_pdf = excel_files["urkal_xlsx"].replace(".xlsx", ".pdf")
    _urkal_to_pdf(excel_files["urkal_xlsx"], urkal_pdf, password)
    pdfs["urkal_pdf"] = urkal_pdf

    if "efb_nachlass_xlsx" in excel_files:
        efb_n_pdf = excel_files["efb_nachlass_xlsx"].replace(".xlsx", ".pdf")
        _efb_to_pdf(excel_files["efb_nachlass_xlsx"], efb_n_pdf, pdf_project)
        pdfs["efb_nachlass_pdf"] = efb_n_pdf

        urkal_n_pdf = excel_files["urkal_nachlass_xlsx"].replace(".xlsx", ".pdf")
        _urkal_to_pdf(excel_files["urkal_nachlass_xlsx"], urkal_n_pdf, password)
        pdfs["urkal_nachlass_pdf"] = urkal_n_pdf
    return pdfs


# ══════════════════════════════════════════════════════════════
#  FORMBLATT 221 GENERATION
# ══════════════════════════════════════════════════════════════

def _fmt(val):
    """German EUR format: 403.137,00"""
    s = f"{val:,.2f}"
    return s.replace(",", "X").replace(".", ",").replace("X", ".")

def _fmt_h(val):
    """German \u20ac/h format: 27,50"""
    return f"{val:.2f}".replace(".", ",")

def _fmt_int(val):
    """German integer with dots: 5.005"""
    s = f"{val:,d}"
    return s.replace(",", ".")

def _build_221_field_values(total, bieter, datum):
    hours = math.ceil(LABOR_RATIO * total / HOURLY_RATE)
    loehne = hours * HOURLY_RATE
    stoffe = round(STOFFE_RATIO * total, 2)
    geraete = round(total - loehne - stoffe, 2)

    return {
        "an_1001": bieter,
        "an_1002": datum,
        "an_01": _fmt_h(27.50), "an_07": "79%", "an_02": _fmt_h(21.73),
        "an_08": "9%", "an_03": _fmt_h(2.48), "an_04": _fmt_h(51.71),
        "an_09": "37,50%", "an_05": _fmt_h(19.39), "an_06": _fmt_h(71.10),
        "an_10": "3,5%", "an_11": "1,5%", "an_12": "0,2%",
        "an_16": "20%", "an_17": "12%", "an_18": "2,8%",
        "an_21": "14%", "an_22": "1,6%", "an_23": "0,5%",
        "an_261": "37,50%", "an_271": "15,10%", "an_281": "3,5%",
        "an_30": _fmt_int(hours), "an_31": _fmt_h(HOURLY_RATE),
        "an_32": _fmt(loehne), "an_35": _fmt(stoffe),
        "an_36": _fmt(geraete), "an_45": _fmt(total),
    }

def _fill_221(template_bytes, total, bieter, datum, output_path):
    field_values = _build_221_field_values(total, bieter, datum)
    pdf = pikepdf.open(io.BytesIO(template_bytes))
    acroform = pdf.Root["/AcroForm"]
    fields = acroform["/Fields"]

    filled = 0
    for field in fields:
        name = str(field.get("/T", ""))
        if name in field_values:
            field["/V"] = pikepdf.String(field_values[name])
            if "/AP" in field:
                del field["/AP"]
            filled += 1

    acroform["/NeedAppearances"] = True
    pdf.save(output_path)
    logger.info(f"221 PDF: filled {filled}/{len(field_values)} fields -> {output_path}")
    return output_path

def generate_221_pdfs(template_bytes, total_original, nachlass_pct, bieter, datum, output_dir):
    files = {}
    orig_path = os.path.join(output_dir, "221_Preisermittlung_Original.pdf")
    _fill_221(template_bytes, total_original, bieter, datum, orig_path)
    files["form221_pdf"] = orig_path

    if nachlass_pct != 0:
        total_nachlass = round(total_original * (1 + nachlass_pct / 100), 2)
        nachlass_label = abs(int(nachlass_pct))
        nachlass_path = os.path.join(output_dir, f"221_Preisermittlung_Nachlass_{nachlass_label}.pdf")
        _fill_221(template_bytes, total_nachlass, bieter, datum, nachlass_path)
        files["form221_nachlass_pdf"] = nachlass_path
    return files
